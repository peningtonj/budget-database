"""
Build a deterministic agency-alias table: for every ingested PBS workbook,
pair the agency's own formal name (extracted from its Table 1.1 "resource
statement" title, which every agency publishes) with the same file's
filename-derived short form (clean_agency() -- the same derivation
program_expenses.agency uses).

Why this instead of fuzzy string matching: both names come from the exact
same file, so the pairing is exact, not a guess. This exists because a
measure's own PBS/PAES sheet identifies a touched agency by its formal
name ("Services Australia") while program_expenses.agency for that same
real agency/year is filename-derived ("SAUS", "SA", ...) -- two
independently-derived spellings of the same real agency, for the same
budget year, from the same single source of PBS documents. Since both
spellings are individually recoverable from that agency's own file, the
correspondence between them is knowable exactly; no heuristic needed.

Cross-YEAR agency drift (the same real agency renamed/restructured over
time) is a different, harder problem, deliberately NOT addressed here --
see program_profile() in backend/measures/views.py, which already joins
on program_name alone rather than guessing agency continuity across
years.
"""
import os
import re
import sqlite3

from openpyxl import load_workbook

from build_db import ROOT, DB_PATH, iter_files, clean_agency, budget_year

TITLE_RE = re.compile(r"^table\s+1\.1:?\s*(.+?)\s+resource\s+statement", re.I)
SHEET_1_1_RE = re.compile(r"(?<![\d.])1\.1(?![\d.])")


def _norm_dashes(s):
    return s.replace("–", "-").replace("—", "-").replace("\xad", " ")


def find_1_1_sheet(sheetnames):
    for name in sheetnames:
        if re.match(r"^program\b", name, re.I):
            continue
        if SHEET_1_1_RE.search(name):
            return name
    return None


def extract_formal_name(path):
    """The agency's own formal name from its Table 1.1 ("Agency Resource
    Statement") title -- every agency publishes this table, unlike Table
    1.2 (measures), which only exists for agencies with measures that
    year. Returns None if no such sheet/title is found (Defence-style
    bespoke layouts, a handful of standards-board agencies, etc. -- these
    remain unresolved rather than guessed at)."""
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return None
    sheet_name = find_1_1_sheet(wb.sheetnames)
    if not sheet_name:
        return None
    ws = wb[sheet_name]
    try:
        title = next(ws.iter_rows(max_row=1))[0].value
    except StopIteration:
        return None
    if not title:
        return None
    m = TITLE_RE.match(_norm_dashes(str(title)).strip())
    if not m:
        return None
    return m.group(1).strip().rstrip("-").strip()


def create_schema(con):
    con.executescript("""
    DROP TABLE IF EXISTS agency_aliases;
    CREATE TABLE agency_aliases (
        id           INTEGER PRIMARY KEY,
        edition      TEXT NOT NULL,
        budget_year  TEXT NOT NULL,
        portfolio    TEXT NOT NULL,
        formal_name  TEXT NOT NULL,   -- from the agency's own Table 1.1 title
        short_name   TEXT NOT NULL,   -- clean_agency(path) -- matches program_expenses.agency
        source_file  TEXT NOT NULL
    );
    CREATE INDEX idx_alias_formal ON agency_aliases(budget_year, formal_name);
    """)


def main():
    con = sqlite3.connect(DB_PATH)
    create_schema(con)

    found = 0
    missed = []
    for edition, portfolio, path in iter_files():
        formal_name = extract_formal_name(path)
        if not formal_name:
            missed.append(os.path.relpath(path, ROOT))
            continue
        con.execute(
            """INSERT INTO agency_aliases
               (edition, budget_year, portfolio, formal_name, short_name, source_file)
               VALUES (?,?,?,?,?,?)""",
            (
                edition,
                budget_year(edition),
                portfolio,
                formal_name,
                clean_agency(path),
                os.path.relpath(path, ROOT),
            ),
        )
        found += 1
    con.commit()

    print(f"Aliases derived : {found}")
    print(f"Files missed    : {len(missed)}")
    for m in missed:
        print("  ", m)
    con.close()


if __name__ == "__main__":
    main()
